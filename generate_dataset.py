#!/usr/bin/env python3
"""
Candidate Dataset Generator
Generates a realistic Excel (.xlsx) talent pool with 120+ diverse candidates.
Uses template-based randomization — no hardcoded candidate entries.
"""

import os
import random
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Data Pools ───────────────────────────────────────────────────────────────

FIRST_NAMES_MALE = [
    "Aarav", "Vivaan", "Aditya", "Vihaan", "Arjun", "Sai", "Reyansh", "Ayaan",
    "Krishna", "Ishaan", "Shaurya", "Atharv", "Advik", "Pranav", "Advaith",
    "Dhruv", "Kabir", "Ritvik", "Aarush", "Kayaan", "Darsh", "Veer", "Sahil",
    "Rohan", "Karan", "Nikhil", "Rahul", "Vikram", "Gaurav", "Manish",
    "Deepak", "Siddharth", "Tushar", "Harsh", "Abhishek", "Suresh", "Rajesh",
    "Anand", "Mohammed", "Faisal", "Zain", "Omar", "Yusuf", "Imran", "Aryan",
]

FIRST_NAMES_FEMALE = [
    "Ananya", "Aanya", "Aadhya", "Aaradhya", "Saanvi", "Prisha", "Anika",
    "Pari", "Angel", "Diya", "Myra", "Sara", "Iraa", "Ahana", "Anvi",
    "Priya", "Sneha", "Kavitha", "Divya", "Meera", "Nisha", "Shruti",
    "Neha", "Pooja", "Riya", "Swathi", "Aparna", "Ishita", "Pallavi",
    "Megha", "Revathi", "Tanvi", "Preeti", "Fatima", "Aisha", "Zara",
    "Gayathri", "Lakshmi", "Sangeetha", "Namrata", "Aishwarya", "Shweta",
]

LAST_NAMES = [
    "Sharma", "Verma", "Patel", "Kumar", "Singh", "Gupta", "Rajan", "Nair",
    "Iyer", "Reddy", "Jain", "Sundaram", "Deshmukh", "Shah", "Mehta",
    "Krishnan", "Menon", "Choudhary", "Mishra", "Saxena", "Malhotra", "Bhatt",
    "Das", "Pillai", "Bhatia", "Narayan", "Kulkarni", "Subramanian", "Pandey",
    "Desai", "Bansal", "Joshi", "Rao", "Tiwari", "Kapoor", "Agarwal",
    "Khan", "Sheikh", "Rizwan", "Prasad", "Venkat", "Raman", "Nambiar",
]

CITIES = [
    "Bangalore", "Hyderabad", "Mumbai", "Delhi", "Pune", "Chennai", "Gurgaon",
    "Noida", "Kolkata", "Ahmedabad", "Kochi", "Jaipur", "Indore", "Chandigarh",
    "Trivandrum", "Coimbatore", "Lucknow", "Nagpur", "Bhopal", "Visakhapatnam",
]

COMPANIES = [
    "Infosys", "TCS", "Wipro", "HCL Technologies", "Tech Mahindra",
    "Cognizant", "Mindtree", "Mphasis", "L&T Infotech", "Persistent Systems",
    "Zoho", "Freshworks", "Razorpay", "Swiggy", "Zomato", "Flipkart",
    "Ola", "PhonePe", "Paytm", "CRED", "Meesho", "Groww", "Zerodha",
    "Dream11", "Postman", "BrowserStack", "Hasura", "Chargebee", "Druva",
    "Unacademy", "BYJU'S", "Lenskart", "Nykaa", "UpGrad", "Simplilearn",
    "Microsoft India", "Google India", "Amazon India", "Goldman Sachs India",
    "JP Morgan India", "Oracle India", "SAP Labs", "Adobe India",
    "Samsung R&D", "Qualcomm India", "Intel India", "IBM India",
    "Accenture", "Deloitte", "EY", "PwC", "KPMG",
    "Startup (Stealth)", "Startup (Series A)", "Startup (Series B)",
    "Freelance / Independent", "Self-employed",
]

EDUCATION = [
    "B.Tech Computer Science", "B.Tech IT", "B.Tech Electronics",
    "B.Tech AI & Data Science", "B.Tech AI", "B.E. Computer Science",
    "B.E. Information Technology", "B.Sc Computer Science", "B.Sc IT",
    "M.Tech Computer Science", "M.Tech AI", "M.Tech AI & ML",
    "M.Tech Software Engineering", "M.Tech NLP", "M.Tech Machine Learning",
    "M.Sc Data Science", "M.Sc Statistics", "M.Sc Mathematics",
    "MS Computer Science", "MBA", "MBA + B.Tech",
    "PhD Computer Science", "PhD Machine Learning", "PhD Statistics",
    "BCA", "MCA", "B.Des Interaction Design",
]

WORK_MODES = ["On-site", "Hybrid", "Remote", "Flexible"]
AVAILABILITY = ["Actively Looking", "Open to Opportunities", "Not Actively Looking", "Available Immediately"]

# ─── Skill Pools by Domain ───────────────────────────────────────────────────

SKILL_POOLS = {
    "ai_ml": {
        "core": ["Python", "TensorFlow", "PyTorch", "Scikit-learn", "Deep Learning", "NLP", "Computer Vision"],
        "llm": ["LLM", "LangChain", "OpenAI API", "RAG", "Vector Databases", "Prompt Engineering", "AI Agents",
                 "Fine-tuning", "Hugging Face", "Transformers"],
        "infra": ["Docker", "AWS", "GCP", "MLOps", "Kubeflow", "AWS SageMaker", "MLflow", "Kubernetes"],
        "data": ["Pandas", "NumPy", "SQL", "Spark", "Feature Engineering", "Statistics", "A/B Testing"],
    },
    "backend": {
        "core": ["Python", "Java", "Go", "Node.js", "C#", "Ruby"],
        "frameworks": ["Django", "FastAPI", "Flask", "Spring Boot", "Express.js", ".NET"],
        "databases": ["PostgreSQL", "MySQL", "MongoDB", "Redis", "Elasticsearch", "Cassandra"],
        "infra": ["Docker", "AWS", "Kubernetes", "CI/CD", "Microservices", "REST APIs", "GraphQL", "gRPC",
                   "Kafka", "RabbitMQ", "Celery"],
    },
    "frontend": {
        "core": ["JavaScript", "TypeScript", "HTML", "CSS"],
        "frameworks": ["React", "Next.js", "Vue.js", "Angular", "Svelte", "Tailwind CSS"],
        "tools": ["Webpack", "Vite", "Jest", "Cypress", "Storybook", "Figma", "GraphQL"],
        "mobile": ["React Native", "Flutter", "Swift", "Kotlin"],
    },
    "data_eng": {
        "core": ["Python", "SQL", "Apache Spark", "Airflow", "dbt"],
        "platforms": ["AWS", "GCP", "Azure", "Snowflake", "BigQuery", "Redshift", "Databricks"],
        "streaming": ["Kafka", "Kinesis", "Flink", "Pulsar"],
        "tools": ["Docker", "Terraform", "CI/CD", "Data Modeling", "ETL", "Data Governance"],
    },
    "devops": {
        "core": ["Docker", "Kubernetes", "Terraform", "Ansible", "CI/CD"],
        "cloud": ["AWS", "GCP", "Azure", "CloudFormation", "Pulumi"],
        "monitoring": ["Prometheus", "Grafana", "ELK Stack", "Datadog", "New Relic"],
        "languages": ["Python", "Bash", "Go", "Ruby"],
        "security": ["IAM", "Vault", "SSL/TLS", "Network Security"],
    },
    "product": {
        "core": ["Product Strategy", "Roadmap Planning", "User Research", "A/B Testing", "Data Analysis"],
        "tools": ["SQL", "Jira", "Figma", "Mixpanel", "Amplitude", "Tableau", "Power BI"],
        "methods": ["Agile", "Scrum", "Design Thinking", "OKRs", "Lean Startup"],
        "technical": ["API Design", "System Architecture", "Python", "JavaScript"],
    },
    "security": {
        "core": ["Network Security", "Penetration Testing", "SIEM", "Incident Response", "Vulnerability Assessment"],
        "tools": ["Burp Suite", "Metasploit", "Wireshark", "Nessus", "OWASP ZAP"],
        "compliance": ["ISO 27001", "SOC 2", "GDPR", "PCI-DSS", "HIPAA"],
        "cloud": ["AWS Security", "Azure Security", "Cloud Security", "IAM", "Zero Trust"],
    },
    "qa": {
        "core": ["Manual Testing", "Automation Testing", "API Testing", "Performance Testing"],
        "tools": ["Selenium", "Cypress", "Playwright", "JMeter", "Postman", "TestNG", "Jest"],
        "languages": ["Python", "Java", "JavaScript", "TypeScript"],
        "methods": ["BDD", "TDD", "CI/CD", "Agile Testing", "Test Planning"],
    },
}

ROLES_BY_DOMAIN = {
    "ai_ml": [
        "AI Engineer", "ML Engineer", "Senior AI Engineer", "Junior AI Engineer",
        "NLP Engineer", "Computer Vision Engineer", "Data Scientist",
        "Senior Data Scientist", "Lead Data Scientist", "AI Researcher",
        "GenAI Engineer", "AI Architect", "MLOps Engineer", "AI Product Engineer",
        "Deep Learning Engineer", "Applied ML Engineer",
    ],
    "backend": [
        "Backend Developer", "Senior Backend Engineer", "Software Engineer",
        "Senior Software Engineer", "Python Developer", "Java Developer",
        "Full Stack Developer", "Technical Lead", "Engineering Manager",
        "Principal Engineer", "Platform Engineer", "API Developer",
    ],
    "frontend": [
        "Frontend Developer", "Senior Frontend Engineer", "UI Engineer",
        "React Developer", "Full Stack Developer", "Frontend Architect",
        "Mobile Developer", "UX Engineer",
    ],
    "data_eng": [
        "Data Engineer", "Senior Data Engineer", "Data Architect",
        "ETL Developer", "Analytics Engineer", "Data Platform Engineer",
    ],
    "devops": [
        "DevOps Engineer", "Senior DevOps Engineer", "SRE",
        "Cloud Engineer", "Infrastructure Engineer", "Platform Engineer",
        "Release Engineer",
    ],
    "product": [
        "Product Manager", "Senior Product Manager", "Associate Product Manager",
        "Technical Product Manager", "Product Owner", "Group Product Manager",
    ],
    "security": [
        "Security Engineer", "Senior Security Engineer", "Security Analyst",
        "Penetration Tester", "Security Architect", "SOC Analyst",
    ],
    "qa": [
        "QA Engineer", "Senior QA Engineer", "SDET", "Test Lead",
        "Automation Engineer", "Performance Test Engineer",
    ],
}

BIO_TEMPLATES = [
    "{role} with {exp} years of experience specializing in {spec}. {achievement}. {interest}.",
    "Experienced {role} focused on {spec}. {achievement}. Currently {interest}.",
    "{exp}-year {role} with expertise in {spec}. {achievement}. {passion}.",
    "Results-driven {role} with a track record in {spec}. {achievement}. {interest}.",
    "Innovative {role} with {exp} years in the industry. {achievement}. {passion}.",
]

ACHIEVEMENTS = [
    "Built and scaled systems serving 1M+ users",
    "Led team of {n} engineers in building core product",
    "Reduced infrastructure costs by {pct}% through optimization",
    "Delivered {n} production-ready features in the last year",
    "Published {n} research papers in top-tier conferences",
    "Improved system performance by {pct}% through re-architecture",
    "Built data pipelines processing {n}TB+ data daily",
    "Deployed {n}+ ML models to production",
    "Contributed to {n}+ open-source projects",
    "Mentored {n} junior developers",
    "Architected microservices handling {n}K+ requests/second",
    "Built RAG system processing {n}K+ documents",
    "Reduced deployment time by {pct}% with CI/CD improvements",
    "Designed API platform used by {n}+ internal teams",
    "Won company hackathon {n} times",
    "AWS/GCP certified professional",
    "Speaker at {n} tech conferences",
    "Built fraud detection system with 99.{n}% accuracy",
]

INTERESTS = [
    "Currently exploring AI/ML and LLM-powered applications",
    "Interested in transitioning to AI engineering",
    "Looking for leadership opportunities in growing teams",
    "Passionate about open-source contributions",
    "Seeking roles with remote-first culture",
    "Interested in startup environments with high growth",
    "Looking to work on large-scale distributed systems",
    "Enthusiastic about building developer tools",
    "Keen on solving real-world problems with technology",
    "Active competitive programmer and Kaggle participant",
    "Interested in EdTech and social impact projects",
    "Looking for roles with strong mentorship culture",
]

PASSIONS = [
    "Passionate about clean code and system design",
    "Driven by building products that impact millions",
    "Loves solving complex engineering challenges",
    "Enthusiastic about the intersection of AI and product",
    "Deeply interested in making technology accessible",
    "Passionate about developer experience and tooling",
]

SPECIALIZATIONS = {
    "ai_ml": ["LLM systems and AI agents", "NLP and transformer architectures",
              "computer vision and deep learning", "MLOps and model deployment",
              "RAG pipelines and semantic search", "production AI/ML systems",
              "generative AI applications", "recommendation systems"],
    "backend": ["high-throughput API systems", "microservices architecture",
                "distributed systems", "event-driven architecture",
                "payment processing systems", "e-commerce platforms",
                "fintech applications", "real-time data processing"],
    "frontend": ["component architecture and design systems", "responsive web applications",
                 "performance optimization", "accessibility and UX",
                 "mobile-first development", "data visualization"],
    "data_eng": ["real-time streaming pipelines", "data warehouse architecture",
                 "ETL/ELT best practices", "big data processing",
                 "data governance and quality", "analytics infrastructure"],
    "devops": ["cloud infrastructure and IaC", "container orchestration",
               "CI/CD pipeline design", "observability and monitoring",
               "cost optimization", "multi-cloud architecture"],
    "product": ["B2B SaaS products", "consumer mobile apps",
                "AI/ML product strategy", "growth and experimentation",
                "fintech products", "healthcare technology"],
    "security": ["application security", "cloud security architecture",
                 "penetration testing and red teaming", "security operations",
                 "compliance and governance", "zero trust architecture"],
    "qa": ["test automation frameworks", "API and integration testing",
           "performance and load testing", "mobile testing",
           "CI/CD testing integration", "security testing"],
}


def _pick_skills(domain: str, exp_years: int) -> list:
    """Pick a realistic skill set based on domain and experience."""
    pool = SKILL_POOLS[domain]
    skills = set()

    # Always pick from core
    core = pool.get("core", [])
    skills.update(random.sample(core, min(len(core), random.randint(2, 4))))

    # Pick from other categories
    other_cats = [k for k in pool if k != "core"]
    for cat in other_cats:
        cat_skills = pool[cat]
        pick_count = random.randint(1, min(3, len(cat_skills)))
        if exp_years >= 5:
            pick_count = min(pick_count + 1, len(cat_skills))
        skills.update(random.sample(cat_skills, pick_count))

    # Senior folks get more skills
    if exp_years >= 7:
        skills.add("System Design")
        if random.random() > 0.5:
            skills.add("Team Leadership")
    if exp_years >= 5 and random.random() > 0.4:
        skills.add("Agile")

    return list(skills)[:random.randint(6, 10)]


def _generate_bio(name: str, role: str, exp: int, domain: str) -> str:
    template = random.choice(BIO_TEMPLATES)
    achievement = random.choice(ACHIEVEMENTS)
    achievement = achievement.format(n=random.randint(2, 15), pct=random.randint(30, 80))
    spec = random.choice(SPECIALIZATIONS.get(domain, ["software development"]))
    interest = random.choice(INTERESTS)
    passion = random.choice(PASSIONS)

    return template.format(
        role=role, exp=exp, spec=spec,
        achievement=achievement, interest=interest, passion=passion
    )


def _salary_for_experience(exp: int, domain: str) -> int:
    """Generate realistic salary expectation (LPA) based on experience and domain."""
    base_map = {
        "ai_ml": {(0, 2): (6, 14), (3, 5): (14, 26), (6, 8): (24, 40), (9, 15): (35, 60)},
        "backend": {(0, 2): (4, 12), (3, 5): (12, 22), (6, 8): (20, 35), (9, 15): (30, 55)},
        "frontend": {(0, 2): (4, 11), (3, 5): (10, 20), (6, 8): (18, 32), (9, 15): (28, 48)},
        "data_eng": {(0, 2): (5, 13), (3, 5): (13, 24), (6, 8): (22, 38), (9, 15): (32, 55)},
        "devops": {(0, 2): (5, 13), (3, 5): (13, 24), (6, 8): (22, 38), (9, 15): (30, 52)},
        "product": {(0, 2): (6, 14), (3, 5): (14, 28), (6, 8): (25, 42), (9, 15): (38, 65)},
        "security": {(0, 2): (5, 12), (3, 5): (12, 24), (6, 8): (22, 38), (9, 15): (32, 55)},
        "qa": {(0, 2): (3, 10), (3, 5): (10, 18), (6, 8): (16, 28), (9, 15): (24, 42)},
    }

    ranges = base_map.get(domain, base_map["backend"])
    for (lo, hi), (sal_lo, sal_hi) in ranges.items():
        if lo <= exp <= hi:
            return random.randint(sal_lo, sal_hi)
    return random.randint(15, 35)


def generate_candidates(count: int = 120) -> list:
    """Generate a list of realistic candidate dictionaries."""
    candidates = []
    used_emails = set()

    # Distribute across domains
    domains = list(ROLES_BY_DOMAIN.keys())
    domain_weights = {
        "ai_ml": 0.25, "backend": 0.20, "frontend": 0.10, "data_eng": 0.12,
        "devops": 0.10, "product": 0.08, "security": 0.08, "qa": 0.07,
    }

    for i in range(count):
        # Pick domain weighted
        domain = random.choices(domains, weights=[domain_weights.get(d, 0.1) for d in domains])[0]

        # Pick name
        is_female = random.random() > 0.45
        first = random.choice(FIRST_NAMES_FEMALE if is_female else FIRST_NAMES_MALE)
        last = random.choice(LAST_NAMES)
        name = f"{first} {last}"

        # Generate unique email
        email_base = f"{first.lower()}.{last[0].lower()}"
        email = f"{email_base}@email.com"
        while email in used_emails:
            email = f"{email_base}{random.randint(1, 99)}@email.com"
        used_emails.add(email)

        # Experience (weighted towards 2-6 years)
        exp = random.choices(
            range(1, 13),
            weights=[3, 8, 12, 15, 15, 12, 10, 8, 6, 4, 3, 2]
        )[0]

        # Pick role appropriate for experience
        role = random.choice(ROLES_BY_DOMAIN[domain])
        if exp <= 2 and "Senior" in role:
            role = role.replace("Senior ", "Junior ")
        elif exp >= 7 and "Junior" in role:
            role = role.replace("Junior ", "Senior ")

        city = random.choice(CITIES)
        company = random.choice(COMPANIES)
        education = random.choice(EDUCATION)
        skills = _pick_skills(domain, exp)
        salary = _salary_for_experience(exp, domain)
        notice = random.choice([0, 15, 15, 30, 30, 30, 45, 45, 60, 60, 90])
        phone = f"+91 {random.randint(70000, 99999)} {random.randint(10000, 99999)}"
        linkedin = f"linkedin.com/in/{first.lower()}{last.lower()}"
        github = f"github.com/{first.lower()}{last[0].lower()}" if random.random() > 0.3 else ""
        work_mode = random.choice(WORK_MODES)
        availability = random.choice(AVAILABILITY)
        bio = _generate_bio(name, role, exp, domain)

        candidates.append({
            "id": f"C{i+1:04d}",
            "name": name,
            "email": email,
            "phone": phone,
            "location": city,
            "experience_years": exp,
            "current_role": role,
            "current_company": company,
            "skills": ", ".join(skills),
            "education": education,
            "salary_expectation_lpa": salary,
            "notice_period_days": notice,
            "linkedin": linkedin,
            "github": github,
            "bio": bio,
            "preferred_work_mode": work_mode,
            "availability_status": availability,
        })

    return candidates


def create_excel(candidates: list, output_path: str):
    """Create a beautifully formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Talent Pool"

    # Column headers
    headers = [
        "ID", "Name", "Email", "Phone", "Location", "Experience (Years)",
        "Current Role", "Current Company", "Skills", "Education",
        "Salary Expectation (LPA)", "Notice Period (Days)", "LinkedIn",
        "GitHub", "Bio", "Preferred Work Mode", "Availability Status"
    ]

    # Styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, cand in enumerate(candidates, 2):
        values = [
            cand["id"], cand["name"], cand["email"], cand["phone"],
            cand["location"], cand["experience_years"], cand["current_role"],
            cand["current_company"], cand["skills"], cand["education"],
            cand["salary_expectation_lpa"], cand["notice_period_days"],
            cand["linkedin"], cand["github"], cand["bio"],
            cand["preferred_work_mode"], cand["availability_status"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [9, 15]))
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-adjust column widths
    col_widths = {
        1: 8, 2: 22, 3: 28, 4: 20, 5: 16, 6: 14, 7: 28, 8: 28,
        9: 50, 10: 30, 11: 18, 12: 16, 13: 30, 14: 25, 15: 60,
        16: 18, 17: 22,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(candidates) + 1}"

    # ─── Summary Sheet ───
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = "4472C4"

    summary_data = [
        ("Talent Pool Summary", ""),
        ("", ""),
        ("Total Candidates", len(candidates)),
        ("Unique Locations", len(set(c["location"] for c in candidates))),
        ("Avg Experience (Years)", round(sum(c["experience_years"] for c in candidates) / len(candidates), 1)),
        ("Avg Salary Expectation (LPA)", round(sum(c["salary_expectation_lpa"] for c in candidates) / len(candidates), 1)),
        ("", ""),
        ("By Experience Level", "Count"),
        ("Junior (1-2 years)", len([c for c in candidates if c["experience_years"] <= 2])),
        ("Mid (3-5 years)", len([c for c in candidates if 3 <= c["experience_years"] <= 5])),
        ("Senior (6-8 years)", len([c for c in candidates if 6 <= c["experience_years"] <= 8])),
        ("Lead (9+ years)", len([c for c in candidates if c["experience_years"] >= 9])),
        ("", ""),
        ("By Location (Top 5)", "Count"),
    ]

    # Top 5 locations
    loc_counts = {}
    for c in candidates:
        loc_counts[c["location"]] = loc_counts.get(c["location"], 0) + 1
    for loc, cnt in sorted(loc_counts.items(), key=lambda x: -x[1])[:5]:
        summary_data.append((loc, cnt))

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(
            name="Calibri", bold=(row_idx in [1, 8, 14] or row_idx <= 1), size=11 if row_idx != 1 else 14
        )
        ws2.cell(row=row_idx, column=2, value=value).font = Font(name="Calibri", size=11)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"[OK] Generated {len(candidates)} candidates -> {output_path}")
    return output_path


if __name__ == "__main__":
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "candidates.xlsx")

    candidates = generate_candidates(120)
    create_excel(candidates, output_path)

    print(f"\nDataset Stats:")
    print(f"   Total: {len(candidates)} candidates")
    print(f"   Locations: {len(set(c['location'] for c in candidates))}")
    print(f"   Avg Experience: {sum(c['experience_years'] for c in candidates) / len(candidates):.1f} years")
    print(f"   Avg Salary: Rs.{sum(c['salary_expectation_lpa'] for c in candidates) / len(candidates):.1f} LPA")
