"""
Flask Web App for Talent Scout Agent
Premium UI with real-time agent progress via Server-Sent Events.
Supports Excel dataset upload, download, and results export.
"""

import os
import sys
import json
import time
import queue
import traceback
import threading
from flask import Flask, render_template, request, jsonify, Response, stream_with_context, send_file
from dotenv import load_dotenv
from werkzeug.utils import secure_filename

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from agent.orchestrator import TalentScoutAgent
from agent.output import ReportGenerator

load_dotenv()

app = Flask(__name__, template_folder="templates", static_folder="static")

# ─── Config ───────────────────────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx"}

# Track the currently active dataset path (default: auto-detect)
_active_dataset_path = None


def _get_active_dataset_path():
    """Get the active dataset file path. Prefers uploaded file, then xlsx, then json."""
    global _active_dataset_path
    if _active_dataset_path and os.path.exists(_active_dataset_path):
        return _active_dataset_path
    
    xlsx_path = os.path.join(DATA_DIR, "candidates.xlsx")
    json_path = os.path.join(DATA_DIR, "candidates.json")
    if os.path.exists(xlsx_path):
        return xlsx_path
    return json_path


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/scout", methods=["POST"])
def scout():
    """Main API endpoint — runs the full agent pipeline."""
    try:
        jd_text = request.form.get("jd_text", "").strip()

        if len(jd_text) < 50:
            return jsonify({"error": "Job Description too short. Minimum 50 characters."}), 400

        agent = TalentScoutAgent()
        result = agent.run(jd_text, pool_path=_get_active_dataset_path())

        if not result.success:
            return jsonify({"error": result.error}), 400

        generator = ReportGenerator()
        json_report = generator.generate_json_report(
            result.ranked_candidates, result.jd_parsed,
            conversations=result.conversations,
            bias_report=result.bias_report,
            market_intel=result.market_intel,
            agent_trace=result.trace
        )

        # Add agent trace to response
        json_report["agent_trace"] = [s.to_dict() for s in result.trace]
        json_report["total_duration_seconds"] = result.total_duration_seconds
        json_report["bias_analysis"] = result.bias_report
        json_report["market_intelligence"] = result.market_intel

        return jsonify(json_report)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/scout/stream", methods=["POST"])
def scout_stream():
    """SSE endpoint — streams agent progress in real-time."""
    jd_text = request.form.get("jd_text", "").strip()

    if len(jd_text) < 50:
        def error_stream():
            yield f"data: {json.dumps({'type': 'error', 'message': 'JD too short'})}\n\n"
        return Response(stream_with_context(error_stream()), mimetype="text/event-stream")

    pool_path = _get_active_dataset_path()
    progress_queue = queue.Queue()

    def on_step(step):
        progress_queue.put({
            "type": "step",
            "step": step.step_number,
            "phase": step.phase,
            "action": step.action,
            "detail": step.detail
        })

    def run_agent_thread():
        try:
            agent = TalentScoutAgent()
            agent.add_callback(on_step)
            result = agent.run(jd_text, pool_path=pool_path)

            if result.success:
                generator = ReportGenerator()
                report = generator.generate_json_report(
                    result.ranked_candidates, result.jd_parsed,
                    conversations=result.conversations,
                    bias_report=result.bias_report,
                    market_intel=result.market_intel,
                    agent_trace=result.trace
                )
                report["agent_trace"] = [s.to_dict() for s in result.trace]
                report["total_duration_seconds"] = result.total_duration_seconds
                report["bias_analysis"] = result.bias_report
                report["market_intelligence"] = result.market_intel
                progress_queue.put({"type": "result", "data": report})
            else:
                progress_queue.put({"type": "error", "message": result.error})
        except Exception as e:
            progress_queue.put({"type": "error", "message": str(e)})
        finally:
            progress_queue.put({"type": "done"})

    thread = threading.Thread(target=run_agent_thread, daemon=True)
    thread.start()

    def event_stream():
        while True:
            try:
                msg = progress_queue.get(timeout=120)
                yield f"data: {json.dumps(msg)}\n\n"
                if msg["type"] in ("done", "error" if "data" not in msg else ""):
                    if msg["type"] == "done":
                        break
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"

    return Response(stream_with_context(event_stream()), mimetype="text/event-stream")


# ─── Dataset Management Endpoints ────────────────────────────────────────────

@app.route("/dataset/info", methods=["GET"])
def dataset_info():
    """Get info about the currently loaded dataset."""
    try:
        from agent.discovery import CandidateDiscovery
        path = _get_active_dataset_path()
        discovery = CandidateDiscovery(pool_path=path)
        stats = discovery.get_pool_stats()
        stats["file_name"] = os.path.basename(path) if path else "None"
        stats["file_size_kb"] = round(os.path.getsize(path) / 1024, 1) if path and os.path.exists(path) else 0
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/dataset/upload", methods=["POST"])
def upload_dataset():
    """Upload a custom Excel dataset to use as the talent pool."""
    global _active_dataset_path

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected."}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": f"Invalid file type '{ext}'. Please upload the .xlsx sample dataset format."}), 400

    # Save with a safe filename
    filename = secure_filename(file.filename)
    save_path = os.path.join(DATA_DIR, filename)
    file.save(save_path)

    # Validate the file can be read
    try:
        from agent.discovery import CandidateDiscovery
        schema_check = CandidateDiscovery.validate_excel_schema(save_path)
        if not schema_check.get("valid"):
            os.remove(save_path)
            return jsonify({"error": schema_check["error"]}), 400

        discovery = CandidateDiscovery(pool_path=save_path)
        stats = discovery.get_pool_stats()
        
        if stats["total_candidates_in_pool"] == 0:
            os.remove(save_path)
            return jsonify({"error": "The uploaded file has the correct columns but no valid candidate rows. Please add candidate data below the header row."}), 400

        # Set as active dataset
        _active_dataset_path = save_path

        stats["file_name"] = filename
        stats["message"] = f"Successfully loaded {stats['total_candidates_in_pool']} candidates from {filename}"
        return jsonify(stats)

    except Exception as e:
        if os.path.exists(save_path):
            os.remove(save_path)
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400


@app.route("/dataset/download-sample", methods=["GET"])
def download_sample_dataset():
    """Download the sample candidates.xlsx dataset."""
    xlsx_path = os.path.join(DATA_DIR, "candidates.xlsx")
    
    if not os.path.exists(xlsx_path):
        # Generate it on the fly
        try:
            from generate_dataset import generate_candidates, create_excel
            candidates = generate_candidates(120)
            create_excel(candidates, xlsx_path)
        except Exception as e:
            return jsonify({"error": f"Could not generate sample dataset: {str(e)}"}), 500

    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name="talent_pool_sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/dataset/reset", methods=["POST"])
def reset_dataset():
    """Reset to the default dataset."""
    global _active_dataset_path
    _active_dataset_path = None
    
    path = _get_active_dataset_path()
    try:
        from agent.discovery import CandidateDiscovery
        discovery = CandidateDiscovery(pool_path=path)
        stats = discovery.get_pool_stats()
        stats["message"] = f"Reset to default dataset: {os.path.basename(path)}"
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/export-results", methods=["POST"])
def export_results():
    """Export agent results to a downloadable Excel file."""
    try:
        data = request.get_json()
        if not data or "candidates" not in data:
            return jsonify({"error": "No results data to export."}), 400

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = Workbook()

        # ─── Sheet 1: Shortlisted Candidates ───
        ws = wb.active
        ws.title = "Shortlisted Candidates"

        headers = [
            "Rank", "Name", "Combined Score", "Match Score", "Interest Score",
            "Recommendation", "Matched Skills", "Missing Skills",
            "Skills Score", "Experience Score", "Salary Score", "Location Score",
            "Availability", "Enthusiasm Signals", "Concerns"
        ]

        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )
        good_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row_idx, cand in enumerate(data["candidates"], 2):
            fb = cand.get("factor_breakdown", {})
            sb = cand.get("skill_breakdown", {})
            values = [
                cand.get("rank", row_idx - 1),
                cand.get("candidate_name", ""),
                cand.get("combined_score", 0),
                cand.get("match_score", 0),
                cand.get("interest_score", 0),
                cand.get("recommendation", ""),
                ", ".join(sb.get("matched_skills", [])[:8]),
                ", ".join(sb.get("missing_skills", [])[:5]),
                f"{fb.get('skills', {}).get('score', 0)}/{fb.get('skills', {}).get('max', 40)}",
                f"{fb.get('experience', {}).get('score', 0)}/{fb.get('experience', {}).get('max', 25)}",
                f"{fb.get('salary', {}).get('score', 0)}/{fb.get('salary', {}).get('max', 20)}",
                f"{fb.get('location', {}).get('score', 0)}/{fb.get('location', {}).get('max', 15)}",
                cand.get("availability_timeline", ""),
                ", ".join(cand.get("enthusiasm_signals", [])[:3]),
                ", ".join(cand.get("concerns", [])[:2]),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 7))
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            # Color code the score
            score = cand.get("combined_score", 0)
            score_cell = ws.cell(row=row_idx, column=3)
            if score >= 85:
                score_cell.fill = good_fill
                score_cell.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
            elif score >= 65:
                score_cell.fill = warn_fill

        # Column widths
        col_widths = [8, 25, 14, 12, 12, 30, 40, 30, 12, 14, 12, 12, 30, 40, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

        # ─── Sheet 2: Summary ───
        ws2 = wb.create_sheet("Summary")
        summary = data.get("summary", {})
        summary_rows = [
            ("Recruitment Agent Results Summary", ""),
            ("", ""),
            ("Report Generated", datetime.now().strftime("%B %d, %Y at %H:%M")),
            ("Job Title", data.get("job_title", "N/A")),
            ("Job Location", data.get("job_location", "N/A")),
            ("", ""),
            ("Total Candidates Evaluated", data.get("total_candidates_considered", 0)),
            ("Priority Hires", summary.get("priority_hires", 0)),
            ("Fast-Track", summary.get("fast_track", 0)),
            ("Recommended", summary.get("recommended", 0)),
            ("Average Match Score", summary.get("average_match_score", 0)),
            ("Average Interest Score", summary.get("average_interest_score", 0)),
            ("Agent Processing Time", f"{data.get('total_duration_seconds', 0)}s"),
        ]

        for row_idx, (label, value) in enumerate(summary_rows, 1):
            ws2.cell(row=row_idx, column=1, value=label).font = Font(
                name="Calibri", bold=(row_idx in [1, 7]), size=14 if row_idx == 1 else 11
            )
            ws2.cell(row=row_idx, column=2, value=value).font = Font(name="Calibri", size=11)

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 25

        # Save to temp file
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_path = os.path.join(OUTPUT_DIR, f"shortlist_{ts}.xlsx")
        wb.save(export_path)

        return send_file(
            export_path,
            as_attachment=True,
            download_name=f"talent_shortlist_{ts}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/health")
def health():
    return jsonify({"status": "ok", "agent": "talent-scout-v3", "version": "3.0"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
