"""
Candidate Discovery Engine
Filters and ranks candidates from a large pool before matching.
Supports both JSON and Excel (.xlsx) data sources.
"""

import json
import os
from typing import List, Dict
from dataclasses import dataclass

# Project root = parent of agent/ directory
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


@dataclass
class DiscoveredCandidate:
    candidate: Dict
    discovery_score: float
    discovery_reason: str


class CandidateDiscovery:
    """
    Discovers candidates from a talent pool stored in Excel or JSON.
    In production, this would query LinkedIn, Naukri, internal DB APIs.
    """
    
    EXPECTED_FIELDS = [
        "id",
        "name",
        "email",
        "phone",
        "location",
        "experience_years",
        "current_role",
        "current_company",
        "skills",
        "education",
        "salary_expectation_lpa",
        "notice_period_days",
        "linkedin",
        "github",
        "bio",
        "preferred_work_mode",
        "availability_status",
    ]

    SAMPLE_HEADERS = [
        "ID",
        "Name",
        "Email",
        "Phone",
        "Location",
        "Experience (Years)",
        "Current Role",
        "Current Company",
        "Skills",
        "Education",
        "Salary Expectation (LPA)",
        "Notice Period (Days)",
        "LinkedIn",
        "GitHub",
        "Bio",
        "Preferred Work Mode",
        "Availability Status",
    ]

    HEADER_MAP = {
        "id": "id",
        "name": "name",
        "email": "email",
        "phone": "phone",
        "location": "location",
        "experience_years": "experience_years",
        "experience": "experience_years",
        "exp_years": "experience_years",
        "current_role": "current_role",
        "role": "current_role",
        "current_company": "current_company",
        "company": "current_company",
        "skills": "skills",
        "education": "education",
        "salary_expectation_lpa": "salary_expectation_lpa",
        "salary_lpa": "salary_expectation_lpa",
        "salary_expectation": "salary_expectation_lpa",
        "salary": "salary_expectation_lpa",
        "notice_period_days": "notice_period_days",
        "notice_period": "notice_period_days",
        "notice_days": "notice_period_days",
        "linkedin": "linkedin",
        "github": "github",
        "bio": "bio",
        "preferred_work_mode": "preferred_work_mode",
        "work_mode": "preferred_work_mode",
        "availability_status": "availability_status",
        "availability": "availability_status",
    }

    def __init__(self, pool_path: str = None):
        if pool_path is None:
            # Prefer Excel over JSON
            xlsx_path = os.path.join(_PROJECT_ROOT, "data", "candidates.xlsx")
            json_path = os.path.join(_PROJECT_ROOT, "data", "candidates.json")
            if os.path.exists(xlsx_path):
                pool_path = xlsx_path
            else:
                pool_path = json_path

        self.pool_path = pool_path
        self.pool = self._load_pool(pool_path)

    @classmethod
    def _normalize_header(cls, header) -> str:
        if header is None:
            return ""
        return (
            str(header).strip().lower()
            .replace(" ", "_")
            .replace("(", "").replace(")", "")
        )

    @classmethod
    def validate_excel_schema(cls, path: str) -> Dict:
        """Validate an uploaded Excel file against the sample dataset schema."""
        ext = os.path.splitext(path)[1].lower()
        if ext != ".xlsx":
            return {
                "valid": False,
                "error": "Please upload an .xlsx file. Download the sample dataset and use the same format."
            }

        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "valid": False,
                "error": "Excel support is not installed. Please install openpyxl and try again."
            }

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return {
                    "valid": False,
                    "error": "This Excel file has no header row. Please look at the sample dataset and update your file accordingly."
                }

            raw_headers = [h for h in header_row if h is not None and str(h).strip()]
            normalized_headers = [cls._normalize_header(h) for h in raw_headers]
            mapped_fields = [cls.HEADER_MAP.get(h) for h in normalized_headers]
            mapped_set = {field for field in mapped_fields if field}

            missing = [field for field in cls.EXPECTED_FIELDS if field not in mapped_set]
            extra = [
                raw_headers[idx]
                for idx, field in enumerate(mapped_fields)
                if field is None
            ]

            if len(raw_headers) < len(cls.EXPECTED_FIELDS) or missing:
                missing_labels = [
                    cls.SAMPLE_HEADERS[cls.EXPECTED_FIELDS.index(field)]
                    for field in missing[:6]
                ]
                suffix = f" Missing columns: {', '.join(missing_labels)}." if missing_labels else ""
                return {
                    "valid": False,
                    "error": "Your dataset has fewer or different columns than expected. Please look at the sample dataset and make the same changes in your file." + suffix
                }

            if len(raw_headers) > len(cls.EXPECTED_FIELDS) or extra:
                extra_labels = ", ".join(str(h) for h in extra[:6]) if extra else "extra columns"
                return {
                    "valid": False,
                    "error": f"Your dataset has more fields than expected ({extra_labels}). Please use only the columns from the sample dataset."
                }

            return {"valid": True}
        finally:
            wb.close()
    
    def _load_pool(self, path: str) -> List[Dict]:
        """Load candidate pool from JSON or Excel file."""
        if not os.path.exists(path):
            print(f"[WARN] Pool file not found: {path}")
            return []

        ext = os.path.splitext(path)[1].lower()

        if ext == ".xlsx":
            return self._load_from_excel(path)
        elif ext == ".json":
            return self._load_from_json(path)
        else:
            print(f"[WARN] Unsupported file format: {ext}")
            return []

    def _load_from_json(self, path: str) -> List[Dict]:
        """Load candidates from a JSON file."""
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_from_excel(self, path: str) -> List[Dict]:
        """
        Load candidates from an Excel (.xlsx) file.
        Converts each row into a candidate dict matching the pipeline's expected format.
        Handles both comma-separated skills strings and list-based skills.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[WARN] openpyxl not installed. Install with: pip install openpyxl")
            return []

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active  # Use the first sheet

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            print("[WARN] Excel file has no data rows.")
            return []

        # Normalize headers to lowercase with underscores
        raw_headers = rows[0]
        headers = [self._normalize_header(h) for h in raw_headers]

        # Map column indices to normalized field names
        col_map = {}
        for idx, header in enumerate(headers):
            mapped = self.HEADER_MAP.get(header)
            if mapped:
                col_map[idx] = mapped

        candidates = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if all(v is None for v in row):
                continue  # Skip empty rows

            cand = {}
            for col_idx, value in enumerate(row):
                field = col_map.get(col_idx)
                if field and value is not None:
                    cand[field] = value

            # Skip if no name
            if not cand.get("name"):
                continue

            # Generate ID if missing
            if not cand.get("id"):
                cand["id"] = f"C{row_idx:04d}"

            # Parse skills: convert comma-separated string to list
            skills_raw = cand.get("skills", "")
            if isinstance(skills_raw, str):
                cand["skills"] = [s.strip() for s in skills_raw.split(",") if s.strip()]
            elif isinstance(skills_raw, list):
                cand["skills"] = skills_raw
            else:
                cand["skills"] = []

            # Ensure numeric fields
            try:
                cand["experience_years"] = int(float(cand.get("experience_years", 0)))
            except (ValueError, TypeError):
                cand["experience_years"] = 0

            try:
                cand["salary_expectation_lpa"] = int(float(cand.get("salary_expectation_lpa", 0)))
            except (ValueError, TypeError):
                cand["salary_expectation_lpa"] = 0

            try:
                cand["notice_period_days"] = int(float(cand.get("notice_period_days", 30)))
            except (ValueError, TypeError):
                cand["notice_period_days"] = 30

            # Defaults for optional fields
            cand.setdefault("location", "Unknown")
            cand.setdefault("current_role", "Software Engineer")
            cand.setdefault("education", "B.Tech")
            cand.setdefault("bio", "")
            cand.setdefault("linkedin", "")
            cand.setdefault("email", "")

            candidates.append(cand)

        wb.close()
        print(f"[OK] Loaded {len(candidates)} candidates from Excel: {os.path.basename(path)}")
        return candidates
    
    def discover(self, jd: Dict, max_results: int = 20) -> List[DiscoveredCandidate]:
        """
        Filter pool by rough relevance, then score by discovery algorithm.
        Returns top candidates worth deep-matching.
        """
        
        required_skills = set(jd.get("required_skills_lower", []))
        nice_skills = set(jd.get("nice_to_have_lower", []))
        jd_location = jd.get("location", "Any").lower()
        exp_min = jd.get("experience_years", {}).get("min") or 0
        exp_max = jd.get("experience_years", {}).get("max") or 10
        
        scored = []
        
        for cand in self.pool:
            score = 0.0
            reasons = []
            
            cand_skills = set([s.lower() for s in cand.get("skills", [])])
            
            # Skill overlap (0-50 points)
            req_match = len(required_skills & cand_skills)
            if required_skills:
                skill_pct = req_match / len(required_skills)
                score += skill_pct * 50
                if req_match >= 2:
                    reasons.append(f"{req_match} required skills match")
            
            # Nice-to-have bonus (0-15 points)
            nice_match = len(nice_skills & cand_skills)
            if nice_skills:
                score += (nice_match / len(nice_skills)) * 15
                if nice_match >= 1:
                    reasons.append(f"{nice_match} nice-to-have skills")
            
            # Experience range (0-20 points)
            cand_exp = cand.get("experience_years", 0)
            if exp_min <= cand_exp <= exp_max:
                score += 20
                reasons.append("Experience in range")
            elif cand_exp >= exp_min * 0.7:
                score += 10
                reasons.append("Experience close to range")
            
            # Location (0-15 points)
            cand_loc = cand.get("location", "").lower()
            if "remote" in jd_location or "any" in jd_location:
                score += 15
            elif jd_location in cand_loc or cand_loc in jd_location:
                score += 15
                reasons.append("Location match")
            else:
                score += 5  # Willing to relocate assumption
            
            if score >= 25:  # Minimum relevance threshold
                scored.append(DiscoveredCandidate(
                    candidate=cand,
                    discovery_score=round(score, 1),
                    discovery_reason="; ".join(reasons) if reasons else "Partial skill match"
                ))
        
        # Sort by discovery score
        scored.sort(key=lambda x: x.discovery_score, reverse=True)
        return scored[:max_results]
    
    def get_pool_stats(self) -> Dict:
        return {
            "total_candidates_in_pool": len(self.pool),
            "locations": list(set(c.get("location", "Unknown") for c in self.pool)),
            "avg_experience": round(sum(c.get("experience_years", 0) for c in self.pool) / len(self.pool), 1) if self.pool else 0,
            "source_file": os.path.basename(self.pool_path) if self.pool_path else "None",
        }
