"""Offline local PySpark ETL for candidate data.

Run: python -m data_pipeline.spark_pipeline
The Flask app never starts Spark. This job creates Parquet plus a JSON adapter
that the existing CandidateDiscovery class can consume as list[dict] data.
"""
import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple

from pyspark.sql import DataFrame, SparkSession
from pyspark.sql import functions as F
from pyspark.sql.types import ArrayType

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DEFAULT_INPUT = DATA_DIR / "candidates.xlsx"
DEFAULT_OUTPUT = DATA_DIR / "processed" / "candidates_parquet"
DEFAULT_ADAPTER = DATA_DIR / "processed" / "candidates.json"

REQUIRED = {"id", "name", "experience_years", "current_role", "skills", "salary_expectation_lpa", "notice_period_days"}
TEXT_COLUMNS = ("id", "name", "email", "phone", "location", "current_role", "current_company", "education", "linkedin", "github", "bio", "preferred_work_mode", "availability_status")
NUMBER_COLUMNS = ("experience_years", "salary_expectation_lpa", "notice_period_days")
HEADER_MAP = {
    "id": "id", "name": "name", "email": "email", "phone": "phone", "location": "location",
    "experience_years": "experience_years", "experience": "experience_years", "exp_years": "experience_years",
    "current_role": "current_role", "role": "current_role", "current_company": "current_company", "company": "current_company",
    "skills": "skills", "education": "education", "salary_expectation_lpa": "salary_expectation_lpa",
    "salary_lpa": "salary_expectation_lpa", "salary_expectation": "salary_expectation_lpa", "salary": "salary_expectation_lpa",
    "notice_period_days": "notice_period_days", "notice_period": "notice_period_days", "notice_days": "notice_period_days",
    "linkedin": "linkedin", "github": "github", "bio": "bio", "preferred_work_mode": "preferred_work_mode",
    "work_mode": "preferred_work_mode", "availability_status": "availability_status", "availability": "availability_status",
}


def get_spark() -> SparkSession:
    # Windows may resolve "python" to the Microsoft Store alias. Point Spark
    # workers at the interpreter running this module instead.
    os.environ["PYSPARK_PYTHON"] = sys.executable
    os.environ["PYSPARK_DRIVER_PYTHON"] = sys.executable
    return (SparkSession.builder.appName("AI-Talent-Scout-Data-Pipeline").master("local[*]")
            .config("spark.sql.session.timeZone", "UTC").getOrCreate())


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("(", "").replace(")", "")


def read_excel_records(path: Path) -> List[Dict]:
    """Excel is read by OpenPyXL, then converted to a real Spark DataFrame."""
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("Excel input has no header row.")
        mapping = {i: HEADER_MAP.get(normalize_header(header)) for i, header in enumerate(headers)}
        records = []
        for row in rows:
            if row and not all(value is None for value in row):
                record = {field: value for i, value in enumerate(row) if (field := mapping.get(i)) and value is not None}
                if record:
                    records.append(record)
        return records
    finally:
        workbook.close()


def ingest(spark: SparkSession, input_path: Path) -> DataFrame:
    if not input_path.exists():
        raise FileNotFoundError(f"Candidate input was not found: {input_path}")
    suffix = input_path.suffix.lower()
    if suffix == ".xlsx":
        records = read_excel_records(input_path)
        if not records:
            raise ValueError("Excel input has no candidate rows.")
        return spark.createDataFrame(records)
    if suffix == ".json":
        return spark.read.option("multiLine", True).json(str(input_path))
    if suffix == ".csv":
        return spark.read.option("header", True).option("inferSchema", True).csv(str(input_path))
    raise ValueError("Supported input formats are .xlsx, .json, and .csv.")


def canonicalize(df: DataFrame) -> DataFrame:
    for column in df.columns:
        target = HEADER_MAP.get(normalize_header(column), normalize_header(column))
        if target != column:
            df = df.withColumnRenamed(column, target)
    return df


def validate(df: DataFrame) -> List[str]:
    if not df.columns or df.limit(1).count() == 0:
        raise ValueError("Candidate dataset is empty.")
    missing = sorted(REQUIRED - set(df.columns))
    if missing:
        raise ValueError("Missing required candidate columns: " + ", ".join(missing))
    warnings = []
    duplicate_ids = df.groupBy("id").count().filter(F.col("count") > 1).count()
    if duplicate_ids:
        warnings.append(f"{duplicate_ids} duplicate ID(s) found; one record per ID will be retained.")
    for column in ("id", "name", "skills"):
        invalid = df.filter(F.col(column).isNull() | (F.trim(F.col(column).cast("string")) == "")).count()
        if invalid:
            warnings.append(f"{invalid} record(s) have an empty {column} value.")
    return warnings


def transform(df: DataFrame) -> DataFrame:
    for column in TEXT_COLUMNS:
        if column in df.columns:
            df = df.withColumn(column, F.trim(F.regexp_replace(F.col(column).cast("string"), r"\s+", " ")))
    for column in NUMBER_COLUMNS:
        df = df.withColumn(column, F.regexp_extract(F.col(column).cast("string"), r"-?\d+(?:\.\d+)?", 0).cast("double").cast("int"))

    skills_type = next(field.dataType for field in df.schema.fields if field.name == "skills")
    if isinstance(skills_type, ArrayType):
        skill_values = F.expr("filter(transform(skills, x -> lower(trim(cast(x as string)))), x -> x <> '')")
    else:
        skill_values = F.expr("filter(transform(split(coalesce(cast(skills as string), ''), ','), x -> lower(trim(x))), x -> x <> '')")
    return (df.withColumn("skills", F.array_distinct(skill_values))
            .fillna({"experience_years": 0, "salary_expectation_lpa": 0, "notice_period_days": 30})
            .dropDuplicates(["id"]))


def write_outputs(df: DataFrame, parquet_path: Path, adapter_path: Path) -> int:
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    adapter_path.parent.mkdir(parents=True, exist_ok=True)
    df.write.mode("overwrite").parquet(str(parquet_path))
    with adapter_path.open("w", encoding="utf-8") as handle:
        json.dump([json.loads(row) for row in df.toJSON().collect()], handle, ensure_ascii=False, indent=2)
    return df.count()


def run_pipeline(input_path: Path = DEFAULT_INPUT, output_path: Path = DEFAULT_OUTPUT, adapter_path: Path = DEFAULT_ADAPTER) -> Tuple[int, List[str]]:
    spark = get_spark()
    try:
        raw = canonicalize(ingest(spark, input_path))
        warnings = validate(raw)
        count = write_outputs(transform(raw), output_path, adapter_path)
        return count, warnings
    finally:
        spark.stop()


def main() -> None:
    parser = argparse.ArgumentParser(description="Build local processed candidate Parquet data.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--adapter", type=Path, default=DEFAULT_ADAPTER)
    args = parser.parse_args()
    count, warnings = run_pipeline(args.input, args.output, args.adapter)
    print(f"[OK] Spark processed {count} candidate(s) into {args.output}")
    print(f"[OK] Agent adapter written to {args.adapter}")
    for warning in warnings:
        print(f"[WARN] {warning}")


if __name__ == "__main__":
    main()

